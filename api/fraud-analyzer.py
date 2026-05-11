"""
RESPACE 부정클릭 자동 분석 API — System 04
Vercel Serverless Function — Python 3.12

변경: pandas/numpy 완전 제거 → 내장 html.parser + openpyxl만 사용 (패키지 ~3MB)
수정: send_response() 순서 버그 교정
"""

from http.server import BaseHTTPRequestHandler
from html.parser import HTMLParser
from collections import defaultdict
from datetime import datetime
from pathlib import Path
import json, re, os, io, base64, tempfile, csv

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── CPC 맵 ───────────────────────────────────────────────────
CPC_MAP = {
    '행사기획': 15500, '행사기획사': 13000, '행사대행사': 13800,
    '팝업스토어제작': 9960, '팝업스토어대행사': 5200, '팝업스토어기획': 5100,
    '팝업컨설팅': 4910, '팝업디자인': 4500, '팝업스토어대행': 4900,
    '팝업대행': 4000, '팝업기획': 3800, '팝업대행사': 5200,
    '팝업스토어시공': 4900, '기업행사기획': 13000, '기업행사전문': 13000,
    '기업행사대행': 13800, '이벤트대행': 5000, '이벤트기획사': 5000,
    'BTL대행사': 5000, '오프라인마케팅': 3000,
}
FOCUS_KW   = set(CPC_MAP.keys())
ADVERTISER = 'https://www.respace.co.kr'

# ── 스타일 상수 ──────────────────────────────────────────────
C_HDR = 'FF1E3A5F'; C_SUB = 'FF2D5A9E'; C_RED = 'FFDC2626'
C_ORG = 'FFF97316'; C_WHITE = 'FFFFFFFF'; C_LGRAY = 'FFF5F5F5'; C_MGRAY = 'FFE0E0E0'

def _fill(c):  return PatternFill('solid', fgColor=c)
def _font(bold=False, color='FF000000', size=10):
    return Font(bold=bold, color=color, size=size, name='맑은 고딕')
def _align(h='center', v='center'):
    return Alignment(horizontal=h, vertical=v, wrap_text=False)
def _border():
    s = Side(border_style='thin', color='FFB0B0B0')
    return Border(left=s, right=s, top=s, bottom=s)
def _row(ws, r, vals, fill=None, bold=False, color='FF000000', h=None):
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.fill = fill or _fill(C_WHITE)
        cell.font = _font(bold=bold, color=color)
        cell.alignment = _align()
        cell.border = _border()
    if h: ws.row_dimensions[r].height = h


# ════════════════════════════════════════════════════════════
# 1. HTML-XLS 파서 (내장 html.parser 사용)
# ════════════════════════════════════════════════════════════
class _TableParser(HTMLParser):
    """에이스카운터 HTML-XLS 파일에서 테이블 데이터를 추출"""
    def __init__(self):
        super().__init__()
        self.tables, self._tbl, self._row, self._cell, self._in = [], None, None, [], False

    def handle_starttag(self, tag, attrs):
        tag = tag.lower()
        if tag == 'table':                   self._tbl = []
        elif tag == 'tr'   and self._tbl is not None: self._row = []
        elif tag in ('td', 'th') and self._row is not None:
            self._cell, self._in = [], True

    def handle_endtag(self, tag):
        tag = tag.lower()
        if tag == 'table' and self._tbl is not None:
            self.tables.append(self._tbl); self._tbl = None
        elif tag == 'tr' and self._row is not None:
            if self._tbl is not None: self._tbl.append(self._row)
            self._row = None
        elif tag in ('td', 'th') and self._in:
            if self._row is not None:
                self._row.append(''.join(self._cell).strip())
            self._cell, self._in = [], False

    def handle_data(self, data):
        if self._in: self._cell.append(data)

    def handle_entityref(self, name):
        if self._in:
            self._cell.append({'nbsp': ' ', 'amp': '&', 'lt': '<', 'gt': '>'}.get(name, ''))

    def handle_charref(self, name):
        if self._in:
            try:
                code = int(name[1:], 16) if name.startswith('x') else int(name)
                self._cell.append(chr(code))
            except Exception:
                pass


def _parse_dt(s):
    """다양한 날짜 포맷 처리"""
    for fmt in ('%Y/%m/%d %H:%M:%S', '%Y-%m-%d %H:%M:%S',
                '%Y/%m/%d %H:%M',    '%Y-%m-%d %H:%M',
                '%Y.%m.%d %H:%M:%S', '%Y.%m.%d %H:%M'):
        try: return datetime.strptime(s.strip(), fmt)
        except ValueError: pass
    return None


def parse_acecounter(filepath):
    """ACE4_15 HTML-XLS → 클릭 레코드 리스트 반환"""
    for enc in ('utf-8', 'euc-kr', 'cp949'):
        try:
            with open(filepath, 'r', encoding=enc, errors='replace') as f:
                raw = f.read()
            break
        except Exception:
            continue
    else:
        raise ValueError("파일을 읽을 수 없습니다.")

    parser = _TableParser()
    parser.feed(raw)

    valid = [t for t in parser.tables
             if len(t) >= 2 and len(t[0]) == 7 and t[0][0].strip() == '순번']

    if not valid:
        # 컬럼 수가 다른 테이블도 확인
        all_lens = [len(t[0]) if t else 0 for t in parser.tables]
        raise ValueError(
            f"파일 형식 오류: ACE4_15 형식이 아닙니다. "
            f"(테이블 {len(parser.tables)}개 감지, 컬럼 수: {set(all_lens)})\n"
            "에이스카운터 > 중복유입 간격분석 > Unique ID 방식으로 "
            "다운로드한 ACE4_15 파일을 사용해 주세요."
        )

    records = []
    for tbl in valid:
        for row in tbl[1:]:  # 헤더 제외
            if not row[0].strip().isdigit():
                continue
            kw     = row[4].strip()
            pv_str = row[6].replace(',', '').strip()
            try:    pv = float(pv_str)
            except: pv = 0.0
            dt = _parse_dt(row[5])
            records.append({
                'IP':        row[1].strip(),
                '광고상품':  row[2].strip(),
                '광고매체':  row[3].strip(),
                '검색어':    kw,
                '유입일시':  row[5].strip(),
                '페이지뷰':  pv,
                '유입일시_dt': dt,
                'CPC':       CPC_MAP.get(kw, 0),
                '집중KW':    kw in FOCUS_KW,
            })

    if not records:
        raise ValueError("파싱된 클릭 레코드가 없습니다. 파일 형식을 확인하세요.")
    return records


# ════════════════════════════════════════════════════════════
# 2. 스코어링 (순수 Python, numpy 없음)
# ════════════════════════════════════════════════════════════
def score_ip(group):
    clicks   = len(group)
    focus    = sum(1 for r in group if r['집중KW'])
    cpc_m    = max((r['CPC'] for r in group), default=0)
    pvs      = [r['페이지뷰'] for r in group]
    avg_pv   = sum(pvs) / len(pvs) if pvs else 0.0

    # 당일 최대 클릭
    by_date = defaultdict(int)
    for r in group:
        if r['유입일시_dt']:
            by_date[r['유입일시_dt'].date()] += 1
    same_day_max = max(by_date.values()) if by_date else 1

    # 최소 유입 간격(분)
    dts = sorted([r['유입일시_dt'] for r in group if r['유입일시_dt']])
    if len(dts) >= 2:
        diffs = [(dts[i+1] - dts[i]).total_seconds() / 60 for i in range(len(dts)-1)]
        pos   = [d for d in diffs if d > 0]
        min_iv = min(pos) if pos else 9999.0
    else:
        min_iv = 9999.0

    # 스코어
    score = 0
    if clicks >= 5:    score += 40
    elif clicks == 4:  score += 30
    elif clicks == 3:  score += 20
    else:              score += 5
    if focus >= 3:     score += 30
    elif focus >= 1:   score += 15
    if cpc_m >= 10000: score += 20
    elif cpc_m >= 5000: score += 10
    if same_day_max >= 3:  score += 25
    elif same_day_max >= 2: score += 15
    if min_iv < 5:     score += 25
    elif min_iv < 60:  score += 10
    if avg_pv >= 10:   score -= 10
    elif avg_pv <= 1.5: score += 10

    if score >= 65:   grade = 'HIGH'
    elif score >= 40: grade = 'MEDIUM-HIGH'
    elif score >= 20: grade = 'MEDIUM'
    else:             grade = 'LOW'

    action = {'HIGH': '즉시 차단', 'MEDIUM-HIGH': '차단 검토',
               'MEDIUM': '모니터링', 'LOW': '정상 관찰'}[grade]

    kws = ' / '.join(list(dict.fromkeys(r['검색어'] for r in group))[:5])
    est_loss = focus * cpc_m if cpc_m > 0 else 0

    if dts:
        first_dt = dts[0].strftime('%Y-%m-%d %H:%M')
        last_dt  = dts[-1].strftime('%Y-%m-%d %H:%M')
        period   = f"{dts[0].strftime('%Y-%m-%d')} ~ {dts[-1].strftime('%Y-%m-%d')}"
    else:
        first_dt = last_dt = period = '-'

    return {
        'IP': group[0]['IP'], '총클릭수': clicks, '집중KW클릭': focus,
        '최고CPC': cpc_m, '예상손실': est_loss, '당일최대': same_day_max,
        '최소간격_분': round(min_iv, 1) if min_iv < 9999 else '-',
        '총PV': int(sum(pvs)), '평균PV': round(avg_pv, 1),
        '위험스코어': score, '위험등급': grade, '권장조치': action,
        '검색어': kws, '유입기간': period, '최초유입': first_dt, '최종유입': last_dt,
    }


def build_ip_list(records):
    grouped = defaultdict(list)
    for r in records:
        grouped[r['IP']].append(r)
    ip_list = [score_ip(g) for g in grouped.values()]
    return sorted(ip_list, key=lambda x: x['위험스코어'], reverse=True)


# ════════════════════════════════════════════════════════════
# 3. Excel 보고서
# ════════════════════════════════════════════════════════════
def _ip_sheet(ws, rows, title, hdr_color):
    cols   = ['IP 주소','총 클릭','집중KW','최고CPC','예상손실','당일최대',
              '최소간격(분)','총PV','평균PV','위험스코어','등급','권장조치']
    keys   = ['IP','총클릭수','집중KW클릭','최고CPC','예상손실','당일최대',
              '최소간격_분','총PV','평균PV','위험스코어','위험등급','권장조치']
    widths = [18,8,8,10,12,8,10,8,8,10,14,12]
    ws.merge_cells(f'A1:{get_column_letter(len(cols))}1')
    ws['A1'].value = title
    ws['A1'].fill, ws['A1'].font, ws['A1'].alignment = _fill(hdr_color), _font(True, C_WHITE, 12), _align()
    ws.row_dimensions[1].height = 26
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    _row(ws, 2, cols, fill=_fill(C_MGRAY), bold=True, h=18)
    gf = {'HIGH': 'FFFEE2E2', 'MEDIUM-HIGH': 'FFFFF3CD', 'MEDIUM': C_LGRAY, 'LOW': C_WHITE}
    for ri, row in enumerate(rows, 3):
        vals = [row.get(k, '') for k in keys]
        if isinstance(vals[3], int) and vals[3] > 0: vals[3] = f'₩{vals[3]:,}'
        if isinstance(vals[4], int) and vals[4] > 0: vals[4] = f'₩{vals[4]:,}'
        _row(ws, ri, vals, fill=_fill(gf.get(row.get('위험등급','LOW'), C_WHITE)), h=16)
    if rows:
        rs = 3 + len(rows)
        total = sum(r['예상손실'] for r in rows)
        ws.cell(row=rs, column=5, value=f'₩{total:,}').fill = _fill(C_MGRAY)


def _detail_sheet(ws, records, ip_dict):
    ws.merge_cells('A1:I1')
    ws['A1'].value = '중복유입 전체 상세 (네이버 클릭 기준)'
    ws['A1'].fill, ws['A1'].font, ws['A1'].alignment = _fill(C_HDR), _font(True, C_WHITE, 12), _align()
    ws.row_dimensions[1].height = 24
    headers = ['클릭일시','키워드','IP','CPC','집중KW','PV','등급','스코어','광고매체']
    for i, w in enumerate([20,18,18,10,8,8,14,8,22], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    _row(ws, 2, headers, fill=_fill(C_MGRAY), bold=True, h=18)
    gf = {'HIGH': 'FFFEE2E2', 'MEDIUM-HIGH': 'FFFFF3CD', 'MEDIUM': C_LGRAY, 'LOW': C_WHITE}
    naver = [r for r in records if '네이버' in r.get('광고상품','')]
    naver.sort(key=lambda r: r['유입일시_dt'] or datetime.min)
    for ri, r in enumerate(naver, 3):
        ip    = r['IP']
        info  = ip_dict.get(ip, {})
        grade = info.get('위험등급', 'LOW')
        dt_s  = r['유입일시_dt'].strftime('%Y-%m-%d %H:%M:%S') if r['유입일시_dt'] else r['유입일시']
        vals  = [dt_s, r['검색어'], ip,
                 f'₩{r["CPC"]:,}' if r['CPC'] > 0 else '-',
                 '●' if r['집중KW'] else '', r['페이지뷰'],
                 grade, info.get('위험스코어', 0), r['광고매체']]
        _row(ws, ri, vals, fill=_fill(gf.get(grade, C_WHITE)), h=16)


def build_excel(ip_list, records, date_tag):
    high   = [r for r in ip_list if r['위험등급'] == 'HIGH']
    med_hi = [r for r in ip_list if r['위험등급'] == 'MEDIUM-HIGH']
    total_l = sum(r['예상손실'] for r in high + med_hi)
    hi_ips  = {r['IP'] for r in high + med_hi}
    naver_cnt = sum(1 for r in records
                    if r['IP'] in hi_ips and '네이버' in r.get('광고상품',''))

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = '요약대시보드'
    for i, w in enumerate([22,18,18,18,18,22], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells('A1:F1')
    ws['A1'].value = f'리스페이스 부정클릭 분석 보고서 ({date_tag})'
    ws['A1'].fill, ws['A1'].font, ws['A1'].alignment = _fill(C_HDR), _font(True, C_WHITE, 13), _align()
    ws.row_dimensions[1].height = 30

    kpi_l = ['총 분석 IP','HIGH (즉시차단)','MED-HIGH (검토)','총 예상 손실액','네이버 클릭 건수']
    kpi_v = [len(ip_list), f'{len(high)}개', f'{len(med_hi)}개', f'₩{total_l:,}', f'{naver_cnt}건']
    ws.row_dimensions[3].height = ws.row_dimensions[4].height = 50
    for col, (lbl, val) in enumerate(zip(kpi_l, kpi_v), 1):
        c = ws.cell(row=3, column=col, value=lbl)
        c.fill, c.font, c.alignment, c.border = _fill(C_HDR), _font(True, C_WHITE, 9), _align(), _border()
        v = ws.cell(row=4, column=col, value=val)
        v.fill = _fill('FFFEF2F2' if col in (2,4) else 'FFEFF6FF')
        v.font, v.alignment, v.border = _font(True, 'FF1E3A5F', 12), _align(), _border()

    ws.merge_cells('A6:F6')
    ws['A6'].value = '등급별 현황'
    ws['A6'].fill, ws['A6'].font, ws['A6'].alignment = _fill(C_SUB), _font(True, C_WHITE), _align()
    ws.row_dimensions[6].height = 22
    _row(ws, 7, ['등급','IP 수','예상 손실액','권장 조치','비율(%)',''], fill=_fill(C_MGRAY), bold=True, h=18)
    total_ips = len(ip_list)
    grade_groups = [
        ('HIGH (즉시차단)',   high,  'FFFEE2E2'),
        ('MEDIUM-HIGH (검토)', med_hi, 'FFFFF3CD'),
        ('MEDIUM (모니터링)', [r for r in ip_list if r['위험등급']=='MEDIUM'], C_LGRAY),
        ('LOW (정상)',        [r for r in ip_list if r['위험등급']=='LOW'],    C_WHITE),
    ]
    for i, (gname, grp, fc) in enumerate(grade_groups, 8):
        cnt  = len(grp)
        loss = sum(r['예상손실'] for r in grp)
        pct  = f'{cnt/total_ips*100:.1f}%' if total_ips else '0%'
        _row(ws, i, [gname, f'{cnt}개', f'₩{loss:,}' if loss else '-', '', pct, ''],
             fill=_fill(fc), h=18)

    ws.merge_cells('A13:F13')
    ws['A13'].value = f'총 예상 손실 합계: ₩{total_l:,}'
    ws['A13'].fill, ws['A13'].font, ws['A13'].alignment = _fill(C_RED), _font(True, C_WHITE, 11), _align()
    ws.row_dimensions[13].height = 24

    _ip_sheet(wb.create_sheet('즉시차단IP(HIGH)'), high, f'즉시차단 HIGH ({date_tag})', C_RED)
    _ip_sheet(wb.create_sheet('차단검토(MED-HI)'), med_hi, f'차단검토 MEDIUM-HIGH ({date_tag})', C_ORG)
    _ip_sheet(wb.create_sheet('전체IP위험도'), ip_list, f'전체 IP 위험도 ({date_tag})', C_SUB)
    ip_dict = {r['IP']: r for r in ip_list}
    _detail_sheet(wb.create_sheet('중복유입전체상세'), records, ip_dict)

    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


# ════════════════════════════════════════════════════════════
# 4. 네이버 SA 접수 CSV
# ════════════════════════════════════════════════════════════
def build_csv(records, ip_list):
    target   = {r['IP'] for r in ip_list if r['위험등급'] in ('HIGH','MEDIUM-HIGH')}
    ip_dict  = {r['IP']: r for r in ip_list}
    naver    = [r for r in records
                if r['IP'] in target and '네이버' in r.get('광고상품','')]
    naver.sort(key=lambda r: r['유입일시_dt'] or datetime.min)

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(['클릭일시','키워드','IP','광고주 URL','부가설명'])
    for r in naver:
        ip   = r['IP']
        info = ip_dict.get(ip, {})
        g    = info.get('위험등급',''); s = info.get('위험스코어', 0)
        t    = '집중KW' if r['집중KW'] else '일반KW'
        note = f"[{g}] 스코어{s}점 | {t} | 동일IP반복클릭"
        dt_s = r['유입일시_dt'].strftime('%Y-%m-%d %H:%M:%S') if r['유입일시_dt'] else r['유입일시']
        writer.writerow([dt_s, r['검색어'], ip, ADVERTISER, note])

    return buf.getvalue().encode('euc-kr', errors='replace'), len(naver)


# ════════════════════════════════════════════════════════════
# 5. Multipart 파싱
# ════════════════════════════════════════════════════════════
def parse_multipart(body: bytes, boundary: str):
    sep   = ('--' + boundary).encode()
    parts = body.split(sep)
    files = {}
    for part in parts[1:]:
        if part in (b'--\r\n', b'--', b''): continue
        split = part.find(b'\r\n\r\n')
        if split == -1: continue
        hdr = part[:split].decode('utf-8', errors='ignore')
        content = part[split + 4:]
        if content.endswith(b'\r\n'): content = content[:-2]
        nm = re.search(r'name="([^"]+)"', hdr)
        fn = re.search(r'filename="([^"]+)"', hdr)
        if nm:
            files[nm.group(1)] = {'content': content, 'filename': fn.group(1) if fn else None}
    return files


# ════════════════════════════════════════════════════════════
# 6. HTTP Handler — send_response() 순서 교정
# ════════════════════════════════════════════════════════════
class handler(BaseHTTPRequestHandler):

    def do_OPTIONS(self):
        self.send_response(200)
        self._cors_headers()
        self.end_headers()

    def do_POST(self):
        try:
            length = int(self.headers.get('Content-Length', 0))
            if length > 20 * 1024 * 1024:
                return self._err(400, '파일 크기는 20MB 이하여야 합니다.')

            body  = self.rfile.read(length)
            ct    = self.headers.get('Content-Type', '')
            bnd_m = re.search(r'boundary=([^\s;]+)', ct)
            if not bnd_m:
                return self._err(400, 'multipart/form-data 형식이 아닙니다.')

            parts = parse_multipart(body, bnd_m.group(1))
            if 'file' not in parts:
                return self._err(400, '파일이 업로드되지 않았습니다.')

            file_data = parts['file']['content']
            filename  = parts['file']['filename'] or 'upload.xls'
            suffix    = '.xlsx' if filename.endswith('.xlsx') else '.xls'

            with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
                tmp.write(file_data); tmp_path = tmp.name

            try:
                records  = parse_acecounter(tmp_path)
                ip_list  = build_ip_list(records)

                high    = [r for r in ip_list if r['위험등급'] == 'HIGH']
                med_hi  = [r for r in ip_list if r['위험등급'] == 'MEDIUM-HIGH']
                med     = [r for r in ip_list if r['위험등급'] == 'MEDIUM']
                low     = [r for r in ip_list if r['위험등급'] == 'LOW']
                total_l = sum(r['예상손실'] for r in high + med_hi)
                hi_ips  = {r['IP'] for r in high + med_hi}
                naver_cnt = sum(1 for r in records
                                if r['IP'] in hi_ips and '네이버' in r.get('광고상품',''))

                m = re.search(r'\((\d{6,8})', filename)
                date_tag = m.group(1)[:4] if m else datetime.now().strftime('%y%m')

                excel_bytes = build_excel(ip_list, records, date_tag)
                excel_name  = f'리스페이스_부정클릭분석_{date_tag}.xlsx'
                csv_bytes, _ = build_csv(records, ip_list)
                csv_name     = f'부정클릭_접수양식_{date_tag}.csv'

            finally:
                os.unlink(tmp_path)

            payload = {
                'filename':    filename,
                'total_count': len(ip_list),
                'high_count':  len(high),
                'medhi_count': len(med_hi),
                'med_count':   len(med),
                'low_count':   len(low),
                'total_loss':  total_l,
                'high_loss':   sum(r['예상손실'] for r in high),
                'medhi_loss':  sum(r['예상손실'] for r in med_hi),
                'naver_count': naver_cnt,
                'excel_b64':   base64.b64encode(excel_bytes).decode(),
                'excel_name':  excel_name,
                'csv_b64':     base64.b64encode(csv_bytes).decode(),
                'csv_name':    csv_name,
            }

            body_out = json.dumps(payload, ensure_ascii=False).encode('utf-8')
            self.send_response(200)
            self._cors_headers()
            self.send_header('Content-Type', 'application/json; charset=utf-8')
            self.send_header('Content-Length', str(len(body_out)))
            self.end_headers()
            self.wfile.write(body_out)

        except ValueError as e:
            self._err(400, str(e))
        except Exception as e:
            self._err(500, f'분석 중 오류가 발생했습니다: {e}')

    def _cors_headers(self):
        """send_response() 호출 이후에만 사용"""
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')

    def _err(self, code, msg):
        body = json.dumps({'error': msg}, ensure_ascii=False).encode('utf-8')
        self.send_response(code)
        self._cors_headers()
        self.send_header('Content-Type', 'application/json; charset=utf-8')
        self.send_header('Content-Length', str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def log_message(self, *args):
        pass
